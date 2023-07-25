import pandas as pd
import random
from tkinter import *
from openpyxl import load_workbook
from playsound import playsound

BRANCO = "#FFFFFF"
PRETO = "#000000"
AZULMEDIO = '#2C6fAB'
AZULESCURO = '#035094'
VERDE = '#00FF00'
VERMELHO = '#FF0000'

abas = ["Esportes", "Games", "História e Geografia", "Arte e Música", "TV e Cultura POP", "Ciências e Engenharia", "Fatos da UFJF"]
ingAbas = ["Sports", "Games", "History and Geography", "Art and Music", "TV and POP Culture", "Science and Engineering"]
opcoes = ["B", "C", "D", "E"]
sons = ["sons/correto.mp3", "sons/errou.mp3"]

def tocaSom(opcao):
    if opcao == 'c':
        playsound(sound= sons[0], block= False)
    elif opcao == 'e':
        playsound(sound= sons[1], block= False)

class functions:
    def __init__ (self, arquivoExc, pergunta, aba, numPergunta, numAba):
        self.arquivoExc = arquivoExc
        self.pergunta = pergunta
        self.aba = aba
        self.numPergunta = numPergunta
        self.numAba = numAba

    def funSorteio (self, dificuldade, txtTema, txtPergunta, botaoA, botaoB, botaoC, botaoD):
        if dificuldade == "N":
            self.numPergunta = random.randint(0, 9)
            self.numAba = random.randint(0, 5)
        elif dificuldade == "D":
            self.numPergunta = random.randint(10, 14)
            self.numAba = random.randint(0, 4)
        elif dificuldade == "E":
            self.numPergunta = random.randint(15, 18)
            self.numAba = random.randint(0, 5)
        elif dificuldade == "F":
            self.numAba = 6
            self.numPergunta = random.randint(0,9)

        self.aba = abas[self.numAba]
        self.arquivoExc = pd.read_excel("planilha/PerguntasX1.xlsx", sheet_name=self.aba)
        self.pergunta = self.arquivoExc.iloc[self.numPergunta, 0]

        if dificuldade == "N" or dificuldade == "D" or dificuldade == "F":
            txtTema["text"] = self.aba
        if dificuldade == "E":
            txtTema["text"] = ingAbas[self.numAba]

        txtPergunta["text"] = self.pergunta

        botaoA["bg"] = AZULMEDIO
        botaoB["bg"] = AZULMEDIO
        botaoC["bg"] = AZULMEDIO
        botaoD["bg"] = AZULMEDIO

        botaoA["fg"] = BRANCO
        botaoB["fg"] = BRANCO
        botaoC["fg"] = BRANCO
        botaoD["fg"] = BRANCO

        botaoA["text"] = self.arquivoExc.iloc[self.numPergunta, 1]
        botaoB["text"] = self.arquivoExc.iloc[self.numPergunta, 2]
        botaoC["text"] = self.arquivoExc.iloc[self.numPergunta, 3]
        botaoD["text"] = self.arquivoExc.iloc[self.numPergunta, 4]

    def funOpcao (self, opcao, botao):
        if opcao == "A":
            planilha = load_workbook("planilha\PerguntasX1.xlsx", data_only = True)
            aba = planilha[abas[self.numAba]]
            cor = aba[f'B{(self.numPergunta)+2}'].fill.start_color.index

            if cor == "FF00FF00" or cor == "FFFF0000":
                botao["bg"] = VERDE
                botao["fg"] = PRETO
                tocaSom("c")
            else:
                botao["bg"] = VERMELHO
                botao["fg"] = PRETO
                tocaSom("e")


        if opcao == "B":
            planilha = load_workbook("planilha\PerguntasX1.xlsx", data_only = True)
            aba = planilha[abas[self.numAba]]
            cor = aba[f'C{(self.numPergunta) + 2}'].fill.start_color.index

            if cor == "FF00FF00" or cor == "FFFF0000":
                botao["bg"] = VERDE
                botao["fg"] = PRETO
                tocaSom("c")
            else:
                botao["bg"] = VERMELHO
                botao["fg"] = PRETO
                tocaSom("e")


        if opcao == "C":
            planilha = load_workbook("planilha\PerguntasX1.xlsx", data_only = True)
            aba = planilha[abas[self.numAba]]
            cor = aba[f'D{(self.numPergunta)+2}'].fill.start_color.index

            if cor == "FF00FF00" or cor == "FFFF0000":
                botao["bg"] = VERDE
                botao["fg"] = PRETO
                tocaSom("c")
            else:
                botao["bg"] = VERMELHO
                botao["fg"] = PRETO
                tocaSom("e")


        if opcao == "D":
            planilha = load_workbook("planilha\PerguntasX1.xlsx", data_only = True)
            aba = planilha[abas[self.numAba]]
            cor = aba[f'E{(self.numPergunta)+2}'].fill.start_color.index

            if cor == "FF00FF00" or cor == "FFFF0000":
                botao["bg"] = VERDE
                botao["fg"] = PRETO
                tocaSom("c")
            else:
                botao["bg"] = VERMELHO
                botao["fg"] = PRETO
                tocaSom("e")
    
    def reset(self, botaoA, botaoB, botaoC, botaoD, txtTema, txtPerg):
        botaoA["text"]=""
        botaoB["text"]=""
        botaoC["text"]=""
        botaoD["text"]=""
        txtTema["text"]=""
        txtPerg["text"]=""
        botaoA["bg"]=AZULMEDIO
        botaoB["bg"]=AZULMEDIO
        botaoC["bg"]=AZULMEDIO
        botaoD["bg"]=AZULMEDIO